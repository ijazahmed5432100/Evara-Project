from django.shortcuts import render,redirect,get_object_or_404
import logging
from django.contrib import messages
from django.contrib.auth import authenticate, login,logout
from django.contrib.auth.decorators import user_passes_test,login_required
from django.contrib.auth.models import User
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.decorators.cache import never_cache
from functools import wraps











import json
# from django.shortcuts import render, redirect,get_object_or_404
# from django.contrib.auth import authenticate, login,logout
# from django.contrib import messages
# from django.contrib.auth.models import User
# from django.contrib.auth.decorators import user_passes_test,login_required
# from django.contrib import messages
# from django.views.decorators.cache import never_cache
# from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Sum, Count, F, DecimalField, Case, When, Value
from django.http import HttpResponse
from django.template.loader import get_template
# from xhtml2pdf import pisa
import csv
from datetime import datetime, timedelta
from products.models import Product, ProductVariant
from coupons.models import Coupon, CouponUsage
from openpyxl import Workbook
from orders.models import Order, OrderItem
from decimal import Decimal
from weasyprint import HTML 
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment  
from openpyxl.utils import get_column_letter
from django.utils import timezone
from datetime import timedelta
from django.utils.dateformat import DateFormat
from django.core.serializers.json import DjangoJSONEncoder
from django.utils import timezone
from django.db.models.functions import TruncHour, TruncDay
import logging
from django.db.models import Min, Sum, Count, Case, When, F, Value
from django.db.models.functions import TruncDay, TruncHour, TruncWeek, TruncMonth




# Create your views here.


"""
ERROR FINDING/DEBUG
"""
logger = logging.getLogger(__name__)






"""
ADMIN CHECK
"""
def is_admin(user):
    return user.is_staff





def admin_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('admin_login')  # your admin login page route name
        if not request.user.is_staff:  # or use a custom user check like user.role == 'admin'
            return redirect('admin_login')  # or maybe a "permission denied" page
        return view_func(request, *args, **kwargs)
    return _wrapped_view






"""
ADMIN LOGIN
"""
@never_cache
def admin_login(request):
    max_attempts = 5
    


    if request.session.get('login_attempts', 0) >= max_attempts:
        messages.error(request, 'Too many failed login attempts. Try again later')
        return render(request, 'admin/admin_login.html')
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')

        user = authenticate(request, username=username,password=password)

        if user is not None and user.is_staff:
            login(request, user)
            request.session['login_attempts'] = 0 #Reset failed attempts
            return redirect('admin_dashboard')
        
        else:
            request.session['login_attempts'] = request.session.get('login_attempts', 0) + 1
            logger.warning(f"Failed admin login attempt for username: {username}")
            messages.error(request, 'Invalid credentials or not an admin user.')
            return render(request, 'admin/admin_login.html')
    
    

    return render(request, 'admin/admin_login.html')


















# from django.shortcuts import render, redirect
# from django.contrib.auth.decorators import login_required, user_passes_test
# from django.utils import timezone
# from django.db.models import Sum, Count
# from django.contrib import messages
# from django.core.paginator import Paginator
# from django.http import HttpResponse
# from django.template.loader import render_to_string
# from weasyprint import HTML
# import pandas as pd
# from io import BytesIO
# from decimal import Decimal
# from orders.models import Order, OrderItem
# from datetime import timedelta
# import json
# import logging
# from django.db.models.functions import TruncDate

# logger = logging.getLogger(__name__)



# @login_required(login_url='admin_login')
# @user_passes_test(is_admin)
# def admin_dashboard(request):
#     # Default date range for the main dashboard (last 30 days)
#     end_date = timezone.now().date()
#     start_date = end_date - timedelta(days=30)
#     date_range = None

#     # Default date range for the graph (earliest order date to today)
#     graph_end_date = timezone.now().date()
#     graph_start_date = Order.objects.aggregate(Min('created_at'))['created_at__min'].date() if Order.objects.exists() else graph_end_date - timedelta(days=30)
#     graph_date_range = None

#     # Handle main dashboard date range filtering (including sales report)
#     if request.method == 'POST' and 'dashboard_filter' in request.POST:
#         date_range = request.POST.get('date_range')
#         custom_start = request.POST.get('custom_start')
#         custom_end = request.POST.get('custom_end')

#         if date_range == 'custom':
#             try:
#                 start_date = timezone.datetime.strptime(custom_start, '%Y-%m-%d').date()
#                 end_date = timezone.datetime.strptime(custom_end, '%Y-%m-%d').date()
#                 if start_date > timezone.now().date():
#                     messages.error(request, "Start date cannot be in the future")
#                     return redirect('admin_dashboard')
#                 if end_date > timezone.now().date():
#                     messages.error(request, "End date cannot be in the future")
#                     return redirect('admin_dashboard')
#                 if start_date > end_date:
#                     messages.error(request, "Start date cannot be after end date")
#                     return redirect('admin_dashboard')
#             except ValueError as e:
#                 messages.error(request, f"Invalid date format: {str(e)}")
#                 return redirect('admin_dashboard')
#             except Exception as e:
#                 messages.error(request, f"An unexpected error occurred: {str(e)}")
#                 return redirect('admin_dashboard')
#         else:
#             end_date = timezone.now().date()
#             if date_range == '1_day':
#                 start_date = end_date - timedelta(days=1)
#             elif date_range == '1_week':
#                 start_date = end_date - timedelta(weeks=1)
#             elif date_range == '1_month':
#                 start_date = end_date - timedelta(days=30)
#             elif date_range == 'yearly':
#                 start_date = end_date.replace(month=1, day=1) - timedelta(days=1)  # Last year start
#                 end_date = end_date.replace(month=12, day=31)  # Last year end
#             else:
#                 messages.error(request, "Invalid date range selected")
#                 return redirect('admin_dashboard')

#     # Handle graph date range filtering
#     if request.method == 'POST' and 'graph_filter' in request.POST:
#         graph_date_range = request.POST.get('graph_date_range')
#         graph_custom_start = request.POST.get('graph_custom_start')
#         graph_custom_end = request.POST.get('graph_custom_end')

#         if graph_date_range == 'custom':
#             try:
#                 graph_start_date = timezone.datetime.strptime(graph_custom_start, '%Y-%m-%d').date()
#                 graph_end_date = timezone.datetime.strptime(graph_custom_end, '%Y-%m-%d').date()
#                 if graph_start_date > timezone.now().date():
#                     messages.error(request, "Graph start date cannot be in the future")
#                     return redirect('admin_dashboard')
#                 if graph_end_date > timezone.now().date():
#                     messages.error(request, "Graph end date cannot be in the future")
#                     return redirect('admin_dashboard')
#                 if graph_start_date > graph_end_date:
#                     messages.error(request, "Graph start date cannot be after end date")
#                     return redirect('admin_dashboard')
#             except ValueError as e:
#                 messages.error(request, f"Invalid graph date format: {str(e)}")
#                 return redirect('admin_dashboard')
#             except Exception as e:
#                 messages.error(request, f"An unexpected error occurred: {str(e)}")
#                 return redirect('admin_dashboard')
#         else:
#             graph_end_date = timezone.now().date()
#             if graph_date_range == '1_day':
#                 graph_start_date = graph_end_date - timedelta(days=1)
#             elif graph_date_range == '1_week':
#                 graph_start_date = graph_end_date - timedelta(weeks=1)
#             elif graph_date_range == '1_month':
#                 graph_start_date = graph_end_date - timedelta(days=30)
#             elif graph_date_range == 'yearly':
#                 graph_start_date = graph_end_date.replace(month=1, day=1) - timedelta(days=1)
#                 graph_end_date = graph_end_date.replace(month=12, day=31)
#             else:
#                 messages.error(request, "Invalid graph date range selected")
#                 return redirect('admin_dashboard')

#     # Get all orders within the main dashboard date range
#     orders = Order.objects.filter(created_at__date__range=[start_date, end_date])
#     completed_orders = orders.filter(status__in=['completed', 'delivered'])

#     # Calculate totals for the main dashboard
#     total_sales = completed_orders.aggregate(total=Sum('total_price'))['total'] or Decimal('0.00')
#     total_discount = completed_orders.aggregate(total=Sum('discount_coupon_amount'))['total'] or Decimal('0.00')
#     total_orders_count = orders.count()

#     # Order status counts
#     order_status_counts = {
#         'pending': orders.filter(status='pending').count(),
#         'completed': completed_orders.count(),
#         'canceled': orders.filter(status='canceled').count(),
#         'returned': orders.filter(status='returned').count(),
#     }

#     # Item status counts
#     item_status_counts = {
#         'delivered': OrderItem.objects.filter(
#             status='delivered',
#             order__created_at__date__range=[start_date, end_date]
#         ).count(),
#         'canceled': OrderItem.objects.filter(
#             status='canceled',
#             order__created_at__date__range=[start_date, end_date]
#         ).count(),
#         'return_requested': OrderItem.objects.filter(
#             status='return_requested',
#             order__created_at__date__range=[start_date, end_date]
#         ).count(),
#         'returned': OrderItem.objects.filter(
#             status='returned',
#             order__created_at__date__range=[start_date, end_date]
#         ).count(),
#     }

#     # Top selling categories
#     top_selling_categories = (
#         OrderItem.objects.filter(
#             order__status__in=['completed', 'delivered'],
#             status__in=['delivered', 'return_requested', 'return_denied'],
#             order__created_at__date__range=[start_date, end_date]
#         )
#         .values('product__category__category_name')
#         .annotate(total_sold=Sum('quantity'))
#         .order_by('-total_sold')[:5]
#     )

#     # Top selling products
#     top_selling_products = OrderItem.objects.filter(
#         order__status__in=['completed', 'delivered'],
#         status__in=['delivered', 'return_requested', 'return_denied'],
#         order__created_at__date__range=[start_date, end_date]
#     ).values('product', 'product__name').annotate(
#         total_sold=Sum('quantity')
#     ).order_by('-total_sold')[:5]

#     # Low stock products (stock < 10)
#     low_stock_products = OrderItem.objects.filter(
#         order__created_at__date__range=[start_date, end_date],
#         order__status__in=['completed', 'delivered']
#     ).values('product_variant').distinct().annotate(
#         stock=Sum('product_variant__stock')
#     ).filter(stock__lt=10)

#     # Prepare data for the sales chart (using graph date range)
#     sales_data = []
#     sales_labels = []
#     orders_data = []

#     graph_completed_orders = Order.objects.filter(
#         status__in=['completed', 'delivered'],
#         created_at__date__range=[graph_start_date, graph_end_date]
#     )

#     sales = graph_completed_orders.annotate(
#         day=TruncDate('created_at')
#     ).values('day').annotate(
#         total_sales=Sum('total_price'),
#         total_orders=Count('id')
#     ).order_by('day')

#     for sale in sales:
#         sales_labels.append(sale['day'].strftime('%Y-%m-%d'))
#         sales_data.append(float(sale['total_sales'] or 0))
#         orders_data.append(sale['total_orders'])

#     # Order conversion data for chart
#     order_conversion_data = {
#         'completed': Order.objects.filter(status__in=['completed', 'delivered'], created_at__date__range=[graph_start_date, graph_end_date]).count(),
#         'canceled': Order.objects.filter(status='canceled', created_at__date__range=[graph_start_date, graph_end_date]).count(),
#         'returned': Order.objects.filter(status='returned', created_at__date__range=[graph_start_date, graph_end_date]).count(),
#         'pending': Order.objects.filter(status='pending', created_at__date__range=[graph_start_date, graph_end_date]).count(),
#     }

#     # Prepare order data for sales report table
#     order_data = []
#     for order in completed_orders:
#         total_before_coupon = sum(item.final_offer_price * item.quantity for item in order.items.all())
#         discounted_amount = order.discount_coupon_amount or Decimal('0.00')
#         order_data.append({
#             'id': order.id,
#             'user': order.user,
#             'total_before_coupon': total_before_coupon,
#             'discounted_amount': discounted_amount,
#             'total_price': order.total_price,
#             'created_at': order.created_at.strftime('%Y-%m-%d %H:%M:%S')
#         })

#     context = {
#         'total_users': 0,  # Placeholder, replace with actual user count if needed
#         'total_sales': total_sales,
#         'total_orders': total_orders_count,
#         'total_discount_amount': total_discount,
#         'order_status_counts': order_status_counts,
#         'item_status_counts': item_status_counts,
#         'top_selling_categories': top_selling_categories,
#         'top_selling_products': top_selling_products,
#         'low_stock_products': low_stock_products,
#         'start_date': start_date.strftime('%Y-%m-%d'),
#         'end_date': end_date.strftime('%Y-%m-%d'),
#         'graph_start_date': graph_start_date.strftime('%Y-%m-%d'),
#         'graph_end_date': graph_end_date.strftime('%Y-%m-%d'),
#         'sales_data_json': json.dumps(sales_data),
#         'sales_labels_json': json.dumps(sales_labels),
#         'orders_data_json': json.dumps(orders_data),
#         'order_conversion_data_json': json.dumps(order_conversion_data),
#         'date_range': date_range,
#         'graph_date_range': graph_date_range,
#         'order_data': order_data,
#         'report_generated_at': timezone.now().strftime('%Y-%m-%d %H:%M:%S')
#     }

#     logger.debug(f"Main Dashboard Start Date: {start_date}, End Date: {end_date}")
#     logger.debug(f"Graph Start Date: {graph_start_date}, End Date: {graph_end_date}")
#     return render(request, 'admin/dashboard.html', context)





# from django.shortcuts import render, redirect
# from django.contrib.auth.decorators import login_required, user_passes_test
# from django.utils import timezone
# from django.db.models import Sum, Count
# from django.contrib import messages
# from django.core.paginator import Paginator
# from django.http import HttpResponse
# from django.template.loader import render_to_string
# from weasyprint import HTML
# import pandas as pd
# from io import BytesIO
# from decimal import Decimal
# from orders.models import Order, OrderItem
# from datetime import timedelta
# import json
# import logging







# # @csrf_exempt
# @login_required(login_url='admin_login')
# @user_passes_test(is_admin)
# def generate_sales_report(request):
#     # Get filter parameters from query string (triggered by generateReport JS)
#     start_date = request.GET.get('start_date')
#     end_date = request.GET.get('end_date')
#     format = request.GET.get('format', 'pdf')

#     if not start_date or not end_date:
#         return HttpResponse("Invalid date range", status=400)

#     try:
#         start_date = timezone.datetime.strptime(start_date, '%Y-%m-%d').date()
#         end_date = timezone.datetime.strptime(end_date, '%Y-%m-%d').date()
#         if start_date > timezone.now().date() or end_date > timezone.now().date() or start_date > end_date:
#             return HttpResponse("Invalid date range", status=400)
#     except ValueError:
#         return HttpResponse("Invalid date format", status=400)

#     # Query completed orders
#     completed_orders = Order.objects.filter(
#         created_at__date__range=[start_date, end_date],
#         status__in=['completed', 'delivered']
#     )
#     total_orders = completed_orders.count()
#     total_sales = completed_orders.aggregate(total=Sum('total_price'))['total'] or Decimal('0.00')
#     total_discount = completed_orders.aggregate(total=Sum('discount_coupon_amount'))['total'] or Decimal('0.00')

#     # Prepare order data for report
#     order_data = []
#     for order in completed_orders:
#         total_before_coupon = sum(item.final_offer_price * item.quantity for item in order.items.all())
#         discounted_amount = order.discount_coupon_amount or Decimal('0.00')
#         order_data.append({
#             'id': order.id,
#             'user': order.user.username,
#             'total_before_coupon': total_before_coupon,
#             'discounted_amount': discounted_amount,
#             'total_price': order.total_price,
#             'created_at': order.created_at.strftime('%Y-%m-%d %H:%M:%S')
#         })

#     context = {
#         'start_date': start_date,
#         'end_date': end_date,
#         'total_orders': total_orders,
#         'total_sales': total_sales,
#         'total_discount_amount': total_discount,
#         'order_status_counts': {
#             'pending': Order.objects.filter(created_at__date__range=[start_date, end_date], status='pending').count(),
#             'completed': completed_orders.count(),
#             'canceled': Order.objects.filter(created_at__date__range=[start_date, end_date], status='canceled').count(),
#             'returned': Order.objects.filter(created_at__date__range=[start_date, end_date], status='returned').count(),
#         },
#         'order_data': order_data,
#         'report_generated_at': timezone.now().strftime('%Y-%m-%d %H:%M:%S')
#     }

#     if format == 'pdf':
#         html_string = render_to_string('admin/sales_report.html', context)
#         pdf = HTML(string=html_string).write_pdf()
#         response = HttpResponse(pdf, content_type='application/pdf')
#         response['Content-Disposition'] = f'attachment; filename=sales_report_{start_date}_to_{end_date}.pdf'
#         return response
#     elif format == 'excel':
#         df = pd.DataFrame(order_data)
#         output = BytesIO()
#         df.to_excel(output, index=False, sheet_name='Sales Report')
#         output.seek(0)
#         response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = f'attachment; filename=sales_report_{start_date}_to_{end_date}.xlsx'
#         return response

#     return HttpResponse("Invalid format", status=400)


















"""
DASHBOARD
"""
@admin_required
@never_cache
@user_passes_test(is_admin)
def admin_dashboard(request):
    
    # Default date range for the main dashboard (last 30 days)
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)
    date_range = None

    # Default date range for the graph (earliest order date to today)
    graph_end_date = timezone.now().date()
    graph_start_date = Order.objects.aggregate(Min('created_at'))['created_at__min'].date() if Order.objects.exists() else graph_end_date - timedelta(days=30)
    graph_date_range = None

    # Handle main dashboard date range filtering
    if request.method == 'POST' and 'dashboard_filter' in request.POST:
        date_range = request.POST.get('date_range')
        custom_start = request.POST.get('custom_start')
        custom_end = request.POST.get('custom_end')

        if date_range == 'custom':
            try:
                start_date = timezone.datetime.strptime(custom_start, '%Y-%m-%d').date()
                end_date = timezone.datetime.strptime(custom_end, '%Y-%m-%d').date()
                
                # Validate dates
                if start_date > timezone.now().date():
                    messages.error(request, "Start date cannot be in the future")
                    return redirect('admin_dashboard')
                if end_date > timezone.now().date():
                    messages.error(request, "End date cannot be in the future")
                    return redirect('admin_dashboard')
                if start_date > end_date:
                    messages.error(request, "Start date cannot be after end date")
                    return redirect('admin_dashboard')
                    
            except ValueError as e:
                messages.error(request, f"Invalid date format: {str(e)}")
                return redirect('admin_dashboard')
            except Exception as e:
                messages.error(request, f"An unexpected error occurred: {str(e)}")
                return redirect('admin_dashboard')
        else:
            end_date = timezone.now().date()
            if date_range == '1_day':
                start_date = end_date - timedelta(days=1)
            elif date_range == '1_week':
                start_date = end_date - timedelta(weeks=1)
            elif date_range == '1_month':
                start_date = end_date - timedelta(days=30)
            else:
                messages.error(request, "Invalid date range selected")
                return redirect('admin_dashboard')

    # Handle graph date range filtering
    if request.method == 'POST' and 'graph_filter' in request.POST:
        graph_date_range = request.POST.get('graph_date_range')
        graph_custom_start = request.POST.get('graph_custom_start')
        graph_custom_end = request.POST.get('graph_custom_end')

        if graph_date_range == 'custom':
            try:
                graph_start_date = timezone.datetime.strptime(graph_custom_start, '%Y-%m-%d').date()
                graph_end_date = timezone.datetime.strptime(graph_custom_end, '%Y-%m-%d').date()
                
                # Validate dates
                if graph_start_date > timezone.now().date():
                    messages.error(request, "Graph start date cannot be in the future")
                    return redirect('admin_dashboard')
                if graph_end_date > timezone.now().date():
                    messages.error(request, "Graph end date cannot be in the future")
                    return redirect('admin_dashboard')
                if graph_start_date > graph_end_date:
                    messages.error(request, "Graph start date cannot be after end date")
                    return redirect('admin_dashboard')
                    
            except ValueError as e:
                messages.error(request, f"Invalid graph date format: {str(e)}")
                return redirect('admin_dashboard')
            except Exception as e:
                messages.error(request, f"An unexpected error occurred: {str(e)}")
                return redirect('admin_dashboard')
        else:
            graph_end_date = timezone.now().date()
            if graph_date_range == '1_day':
                graph_start_date = graph_end_date - timedelta(days=1)
            elif graph_date_range == '1_week':
                graph_start_date = graph_end_date - timedelta(weeks=1)
            elif graph_date_range == '1_month':
                graph_start_date = graph_end_date - timedelta(days=30)
            else:
                messages.error(request, "Invalid graph date range selected")
                return redirect('admin_dashboard')

    # Get all orders within the main dashboard date range
    orders = Order.objects.filter(created_at__date__range=[start_date, end_date])

    # Calculate totals for the main dashboard
    total_sales = Decimal('0.00')
    total_discount = Decimal('0.00')
    
    for order in orders.filter(status='completed'):
        valid_items = order.items.filter(
            status__in=['delivered', 'return_requested', 'return_denied']
        )
        order_total = sum(Decimal(str(item.price)) * Decimal(str(item.quantity)) for item in valid_items)
        total_sales += order_total
        
        if order.coupon:
            total_discount += Decimal(str(order.balance_refund))

    # Other dashboard data
    total_users = User.objects.filter(is_superuser=False).count()
    total_orders_count = orders.count()  # Total orders (pending + completed + canceled + returned)

    order_status_counts = {
        'pending': orders.filter(status='pending').count(),
        'completed': orders.filter(status='completed').count(),
        'canceled': orders.filter(status='canceled').count(),
        'returned': orders.filter(status='returned').count(),
    }

    item_status_counts = {
        'delivered': OrderItem.objects.filter(
            status='delivered',
            order__created_at__date__range=[start_date, end_date]
        ).count(),
        'canceled': OrderItem.objects.filter(
            status='canceled',
            order__created_at__date__range=[start_date, end_date]
        ).count(),
        'return_requested': OrderItem.objects.filter(
            status='return_requested',
            order__created_at__date__range=[start_date, end_date]
        ).count(),
        'returned': OrderItem.objects.filter(
            status='returned',
            order__created_at__date__range=[start_date, end_date]
        ).count(),
    }

    top_selling_categories = (
        OrderItem.objects.filter(
            order__status='completed',
            status__in=['delivered', 'return_requested', 'return_denied'],
            order__created_at__date__range=[start_date, end_date]
        )
        .values('product__category__category_name')
        .annotate(total_sold=Sum('quantity'))
        .order_by('-total_sold')[:5]
    )

    top_selling_products = Product.objects.annotate(
        total_sold=Sum(
            Case(
                When(
                    variants__orderitem__order__status='completed',
                    variants__orderitem__status__in=['delivered', 'return_requested', 'return_denied'],
                    variants__orderitem__order__created_at__date__range=[start_date, end_date],
                    then=F('variants__orderitem__quantity')
                ),
                default=Value(0),
                output_field=DecimalField()
            )
        )
    ).order_by('-total_sold')[:5]

    low_stock_products = ProductVariant.objects.filter(stock__lt=10)

    # Prepare data for the sales chart (using graph date range)
    sales_data = []
    sales_labels = []
    orders_data = []

    graph_completed_orders = Order.objects.filter(
        status='completed',
        created_at__date__range=[graph_start_date, graph_end_date]
    )

    # Always aggregate by day, regardless of the selected filter
    sales = graph_completed_orders.annotate(day=TruncDay('created_at')).values('day').annotate(
        total_sales=Sum('total_price'),
        total_orders=Count('id')
    ).order_by('day')

    for sale in sales:
        sales_labels.append(sale['day'].strftime('%Y-%m-%d'))
        sales_data.append(float(sale['total_sales']))
        orders_data.append(sale['total_orders'])

    # Prepare data for the circular graph (order conversion analysis)
    order_conversion_data = {
        'completed': Order.objects.filter(status='completed', created_at__date__range=[graph_start_date, graph_end_date]).count(),
        'canceled': Order.objects.filter(status='canceled', created_at__date__range=[graph_start_date, graph_end_date]).count(),
        'returned': Order.objects.filter(status='returned', created_at__date__range=[graph_start_date, graph_end_date]).count(),
        'pending': Order.objects.filter(status='pending', created_at__date__range=[graph_start_date, graph_end_date]).count(),
    }

    
    recent_orders = Order.objects.order_by('-created_at')[:10]  # Last 10 orders

    context = {
        'total_users': total_users,
        'total_sales': total_sales - total_discount,
        'total_orders': total_orders_count,
        'total_discount': total_discount,
        'order_status_counts': order_status_counts,
        'item_status_counts': item_status_counts,
        'top_selling_categories': top_selling_categories,
        'top_selling_products': top_selling_products,
        'low_stock_products': low_stock_products,
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'graph_start_date': graph_start_date.strftime('%Y-%m-%d'),
        'graph_end_date': graph_end_date.strftime('%Y-%m-%d'),
        'sales_data_json': json.dumps(sales_data, cls=DjangoJSONEncoder),
        'sales_labels_json': json.dumps(sales_labels, cls=DjangoJSONEncoder),
        'orders_data_json': json.dumps(orders_data, cls=DjangoJSONEncoder),
        'order_conversion_data_json': json.dumps(order_conversion_data, cls=DjangoJSONEncoder),
        'date_range': date_range,  # For main dashboard filter
        'graph_date_range': graph_date_range,  # For graph filter
        'recent_orders': recent_orders,
    }
    
    logger.debug(f"Main Dashboard Start Date: {start_date}, End Date: {end_date}")
    logger.debug(f"Graph Start Date: {graph_start_date}, End Date: {graph_end_date}")
    return render(request, 'admin/dashboard.html', context)













"""
SALES REPORT GENERATE
"""
@admin_required
@never_cache
@user_passes_test(is_admin)
def generate_sales_report(request):
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    report_format = request.GET.get('format', 'pdf')

    # Validate that start_date and end_date are provided
    if not start_date_str or not end_date_str:
        return HttpResponse("Start date and end date are required", status=400)

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        # Validate dates
        if start_date > end_date:
            return HttpResponse("Start date cannot be after end date", status=400)
        if end_date > timezone.now().date():
            return HttpResponse("End date cannot be in the future", status=400)
            
    except ValueError:
        return HttpResponse("Invalid date format", status=400)

    # Get all orders within the date range
    orders = Order.objects.filter(created_at__date__range=[start_date, end_date])

    # Calculate totals for the report
    total_sales = Decimal('0.00')
    total_discount = Decimal('0.00')
    order_data = []

    for order in orders.filter(status='completed'):
        valid_items = order.items.filter(
            status__in=['delivered', 'return_requested', 'return_denied']
        )
        
        order_total = sum(item.price * item.quantity for item in valid_items)
        order_discount = order.balance_refund if order.coupon else Decimal('0.00')
        
        order_data.append({
            'id': order.id,
            'user': order.user,
            'total_before_coupon': order_total,
            'total_price': order_total  - order_discount,
            'discounted_amount': order_discount,
            'created_at': order.created_at
        })
        
        total_sales += order_total
        total_discount += order_discount

    # Calculate order status counts
    order_status_counts = {
        'pending': orders.filter(status='pending').count(),
        'completed': orders.filter(status='completed').count(),
        'canceled': orders.filter(status='canceled').count(),
        'returned': orders.filter(status='returned').count(),
    }

    context = {
        'order_data': order_data,
        'total_sales': total_sales - total_discount,
        'total_discount_amount': total_discount,
        'total_orders': orders.count(),  # Total orders (pending + completed + canceled + returned)
        'order_status_counts': order_status_counts,  # Include order status counts
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'report_generated_at': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
    }

    if report_format == 'pdf':
        template = get_template('admin/sales_report.html')
        html = template.render(context)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="sales_report_{start_date}_{end_date}.pdf"'
        HTML(string=html).write_pdf(response)
        return response
    elif report_format == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="sales_report_{start_date}_{end_date}.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Sales Report'

        # Define styles
        title_font = Font(bold=True, size=16, name='Arial', color='2563EB')
        header_font = Font(bold=True, color="FFFFFF", name='Arial', size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        company_font = Font(bold=True, size=12, name='Arial', color='1e40af')
        address_font = Font(size=10, name='Arial', color='666666')
        summary_title_font = Font(bold=True, size=12, name='Arial', color='1e40af')
        summary_label_font = Font(bold=True, size=10, name='Arial', color='4b5563')
        summary_value_font = Font(bold=True, size=11, name='Arial', color='1e40af')
        normal_font = Font(name='Arial', size=10)
        amount_font = Font(name='Consolas', size=10, color='1e40af') 

        border = Border(
            left=Side(style='thin', color='e5e7eb'),
            right=Side(style='thin', color='e5e7eb'),
            top=Side(style='thin', color='e5e7eb'),
            bottom=Side(style='thin', color='e5e7eb')
        )

        # Alignments
        center_align = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        # Background fills
        light_gray_fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
        summary_item_fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")

        # Title Section
        worksheet.merge_cells('A1:F1')
        title_cell = worksheet.cell(row=1, column=1, value="SALES REPORT")
        title_cell.font = title_font
        title_cell.alignment = center_align
        
        # Company Info Section
        worksheet.merge_cells('A3:F3')
        company_cell = worksheet.cell(row=3, column=1, value="URBAN EDGE")
        company_cell.font = company_font
        company_cell.alignment = left_align

        # Company Address
        address_text = "National Highway 66 near Calicut University\nKakkanchery Chelembra PO, Dt, Thenhipalam\nKerala 673634"
        worksheet.merge_cells('A4:F4')
        address_cell = worksheet.cell(row=4, column=1, value=address_text)
        address_cell.font = address_font
        address_cell.alignment = left_align

        # Period Info
        worksheet.merge_cells('A6:F6')
        period_cell = worksheet.cell(row=6, column=1, value=f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
        period_cell.font = normal_font
        period_cell.alignment = left_align

        # Summary Section
        worksheet.merge_cells('A8:F8')
        summary_title = worksheet.cell(row=8, column=1, value="Summary")
        summary_title.font = summary_title_font

        #Adjesting the sales amount according to the discount
        total_sales -= total_discount

        # Summary Data
        summary_data = [
            #('Total Orders', orders.count()),
            ('Total Sales', f"₹{total_sales:,.2f}"),
            ('Total Discount', f"₹{total_discount:,.2f}")
        ]

        # Create summary grid
        start_row = 9
        for idx, (label, value) in enumerate(summary_data):
            # Labels
            cell = worksheet.cell(row=start_row + idx, column=1, value=label)
            cell.font = summary_label_font
            cell.fill = summary_item_fill
            cell.border = border
            cell.alignment = left_align
            
            # Values
            cell = worksheet.cell(row=start_row + idx, column=2, value=value)
            cell.font = summary_value_font
            cell.fill = summary_item_fill
            cell.border = border
            cell.alignment = right_align

        # Order Status Counts Section
        status_row = start_row + len(summary_data) + 2
        worksheet.merge_cells(f'A{status_row}:F{status_row}')
        status_title = worksheet.cell(row=status_row, column=1, value="Order Status Counts")
        status_title.font = summary_title_font

        # Order Status Data
        status_data = [
            ('Total Orders', orders.count()),
            ('Pending', order_status_counts['pending']),
            ('Completed', order_status_counts['completed']),
            ('Canceled', order_status_counts['canceled']),
            ('Returned', order_status_counts['returned']),
        ]

        # Create status grid
        for idx, (label, value) in enumerate(status_data):
            # Labels
            cell = worksheet.cell(row=status_row + idx + 1, column=1, value=label)
            cell.font = summary_label_font
            cell.fill = summary_item_fill
            cell.border = border
            cell.alignment = left_align
            
            # Values
            cell = worksheet.cell(row=status_row + idx + 1, column=2, value=value)
            cell.font = summary_value_font
            cell.fill = summary_item_fill
            cell.border = border
            cell.alignment = right_align

        # Order Details Section
        detail_row = status_row + len(status_data) + 2
        worksheet.merge_cells(f'A{detail_row}:F{detail_row}')
        detail_title = worksheet.cell(row=detail_row, column=1, value="Order Details")
        detail_title.font = summary_title_font

        # Headers
        headers = ['Order ID', 'Customer', 'Amount (₹)', 'Discount (₹)', 'After Discount (₹)', 'Date']
        header_row = detail_row + 1
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align

        # Data rows
        for idx, order in enumerate(order_data, 1):
            row = header_row + idx
            data = [
                order['id'],
                order['user'].username,
                f"₹{order['total_before_coupon']:,.2f}",
                f"₹{order['discounted_amount']:,.2f}",
                f"₹{order['total_price']:,.2f}",
                order['created_at'].strftime('%Y-%m-%d %H:%M:%S')
            ]
            
            for col, value in enumerate(data, 1):
                cell = worksheet.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = right_align if isinstance(value, str) and '₹' in value else center_align
                cell.font = amount_font if isinstance(value, str) and '₹' in value else normal_font
                if row % 2 == 0:
                    cell.fill = light_gray_fill

        # Set column widths
        column_widths = [25, 20, 18, 18, 18, 22]
        for i, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = width

        # Generated date at bottom
        last_row = worksheet.max_row + 2
        worksheet.merge_cells(f'A{last_row}:F{last_row}')
        generated_cell = worksheet.cell(row=last_row, column=1, value=f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}")
        generated_cell.font = normal_font
        generated_cell.alignment = center_align

        workbook.save(response)
        return response
    
    return redirect(request, 'admin_dashboard')















"""
ADMIN SIDE USER MANAGEMENT
"""
@admin_required
@never_cache
@user_passes_test(is_admin)
def user_manage(request):
    query = request.GET.get('q', '') 
    users = User.objects.filter(is_superuser=False).order_by('-date_joined')

    if query:
        users = users.filter(username__icontains=query)
        

    page = request.GET.get('page', 1)
    paginator = Paginator(users, 10)


    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)        

    return render(request, 'admin/users.html', {'users':users , 'query': query,})





"""
BLOCK USER
"""
@admin_required
@never_cache
@user_passes_test(is_admin)
def block_user(request,user_id):
    user = get_object_or_404(User, id=user_id)
    user.is_active = False
    user.save()
    messages.success(request, f"{user.username} has been blocked.")
    return redirect('users')






"""
UNBLOCK USER
"""
@admin_required
@never_cache
@user_passes_test(is_admin)
def unblock_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.is_active = True
    user.save()
    messages.success(request, f'{user.username} has been unblocked.')
    return redirect('users')






"""
ADMIN LOGOUT
"""
@admin_required
@never_cache
@user_passes_test(is_admin)
def admin_logout(request):
    logout(request)
    return redirect('admin_login')














